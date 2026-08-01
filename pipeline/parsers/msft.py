"""Microsoft parser.

Unlike the other four companies, MSFT has no press release in this dataset -
all structured financials come from 'Microsoft statements.xlsx', Microsoft's
standard quarterly IR workbook. As of this dataset, the workbook's per-quarter
sheets ('Segment Results', 'Cash Flows') and the 'Quarterly Income Statements'
history sheet are populated only through Q1 FY26; Q2 FY26 columns exist as
empty placeholders (headers with no data), and the only Q2 artifact is a
5-slide, image-only 'Outlook' deck with no extractable text. So Q2 is
genuinely "not available" for Microsoft in this dataset - this parser only
ever produces a Q1 record.
"""
CURRENCY = "USD"

_SEGMENT_ROWS = {
    "Productivity and Business Processes": {"revenue": 10, "cost_of_revenue": 11, "opex": 12, "operating_income": 13},
    "Intelligent Cloud": {"revenue": 15, "cost_of_revenue": 16, "opex": 17, "operating_income": 18},
    "More Personal Computing": {"revenue": 20, "cost_of_revenue": 21, "opex": 22, "operating_income": 23},
}


def _find_quarter_column(ws, header_row: int, label: str):
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=header_row, column=col).value == label:
            return col
    return None


def parse_workbook(path) -> dict:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)

    # --- Quarterly Income Statements: wide history sheet, Q1-26 is the last populated column ---
    qis = wb["Quarterly Income Statements"]
    col = _find_quarter_column(qis, header_row=5, label="Q1-26")
    kpis = {"currency": CURRENCY}
    if col is not None:
        def cell(row):
            return qis.cell(row=row, column=col).value

        revenue = cell(9)
        cost_of_revenue = cell(13)
        gross_margin = cell(14)
        op_income = cell(19)
        net_income = cell(23)
        diluted_eps = cell(26)

        kpis.update({
            "revenue": revenue,
            "cost_of_revenue": cost_of_revenue,
            "gross_profit": gross_margin,
            "gross_margin_pct": round(gross_margin / revenue * 100, 1) if revenue and gross_margin is not None else None,
            "operating_income": op_income,
            "operating_margin_pct": round(op_income / revenue * 100, 1) if revenue and op_income is not None else None,
            "net_income": net_income,
            "diluted_eps": diluted_eps,
        })

    # --- Cash Flows: single-quarter sheet (col B = current quarter = Q1 FY26) ---
    cf = wb["Cash Flows"]
    ocf = cf.cell(row=26, column=2).value
    capex = cf.cell(row=36, column=2).value
    kpis["operating_cash_flow"] = ocf
    kpis["capex"] = capex
    if ocf is not None and capex is not None:
        kpis["free_cash_flow"] = round(ocf + capex, 1)  # capex already negative in this sheet

    # --- Balance Sheets: single-quarter sheet (col B = current quarter = Q1 FY26) ---
    bs = wb["Balance Sheets"]
    cash_and_st_investments = bs.cell(row=11, column=2).value
    current_debt = bs.cell(row=26, column=2).value
    long_term_debt = bs.cell(row=32, column=2).value
    if cash_and_st_investments is not None and current_debt is not None and long_term_debt is not None:
        kpis["net_debt"] = round((current_debt + long_term_debt) - cash_and_st_investments, 1)

    # --- Segment Results: single-quarter sheet (col B = current quarter = Q1 FY26) ---
    seg = wb["Segment Results"]
    business_units = []
    for name, rows in _SEGMENT_ROWS.items():
        revenue = seg.cell(row=rows["revenue"], column=2).value
        op_income_bu = seg.cell(row=rows["operating_income"], column=2).value
        if revenue is not None:
            business_units.append({
                "name": name,
                "revenue": revenue,
                "operating_income": op_income_bu,
                "operating_margin_pct": round(op_income_bu / revenue * 100, 1) if revenue and op_income_bu is not None else None,
                "currency": CURRENCY,
            })

    return {
        "quarter": "Q1",
        "kpis": kpis,
        "business_units": business_units,
        "guidance": {},
        "commentary": [],
    }
